from django.contrib.auth import authenticate, login
from django.contrib.auth.models import User
from django.conf import settings
from django.http import HttpResponsePermanentRedirect
from django.shortcuts import render, reverse

from .models import Book, Author, BookInstance, Status, TextbookInstance, Textbook, Genre, PublishingHouse, \
    StudentGroup, UserData, IssueTextbooks
from .forms import RegisterForm, LoginForm, AddNewBookForm, ChangePasswordForm, \
    AddTextBookFromExcelForm, AddNewBookInstanceForm, IssueTextbookForm, IssueABookForm


from .services import *
import openpyxl
import datetime
import logging

# подключение файла для логирования
logging.basicConfig(filename='site_logging.log',
                    format="%(asctime)s : %(levelname)s - %(funcName)s: %(lineno)d - %(message)s",
                    level=logging.INFO)


# главная страница
def index(request):
    return render(request, 'index.html')


# страница с информацией о сайте
def about(request):
    return render(request, 'about.html')


'''   Book Views   '''


def catalog(request):
    page = request.GET.get('page', 1)
    genre = request.GET.get('genre', None)
    genres = Genre.objects.all()

    if genre:
        genre_label = genres.filter(id=genre)[0].title
        page = get_books_with_pagination(page=page, genres=[genre])
    else:
        genre_label = 'Все книги'
        page = get_books_with_pagination(page=page)
    context = {'page': page,
               'genres': genres,
               'genre_label': genre_label,
               'media': settings.STATIC_URL}
    return render(request, 'catalog.html', context=context)


def book_item(request, book_id):
    book = Book.objects.get(id=book_id)
    count = book.bookinstance_set.all().filter(status=STATUS_FREE).count()
    user_from_request = request.user
    try:
        user_books = BookInstance.objects.get(borrower=user_from_request)
        have_book = bool(user_books)
    except Exception:
        have_book = False
    return render(request, 'book_item.html', context={'book': book,
                                                      'count': count,
                                                      'have_book': have_book})


'''   Author Views   '''


def authors_view(request):
    authors = Author.objects.all()
    context = {'authors': authors, 'media': settings.STATIC_URL}
    return render(request, 'authors.html', context=context)


def author_person_view(request, author_id):
    author = Author.objects.get(id=author_id)
    books = author.book.all()
    context = {'author': author, 'books': books}
    return render(request, 'author_person.html', context=context)


'''   User Views   '''


def register(request):
    if request.method == 'POST':
        register_form = RegisterForm(request.POST)

        if register_form.is_valid():
            # получение данных из формы
            username = request.POST.get('username')
            first_name = request.POST.get('name')
            last_name = request.POST.get('last_name')
            e_mail = request.POST.get('e_mail')
            password = request.POST.get('password')
            group = request.POST.get('group')
            try:
                # пробуем добавить нового юзера
                level = int(group[:-1])
                letter = group[-1].upper()
                create_user(username, e_mail, password, first_name, last_name, (level, letter))
            except Exception as e:
                # если данные некорректны, сообщаем об этом пользователю
                print(e)
                register_form.add_error('group', f'Класс {group} отсутствует в базе')
                return render(request, 'register.html', {'form': register_form})
            return HttpResponsePermanentRedirect(reverse('login_page'))
    else:
        register_form = RegisterForm()
        return render(request, 'register.html', {'form': register_form})


def login_view(request):
    if request.method == 'POST':
        login_form = LoginForm(request.POST)

        if login_form.is_valid():
            username = request.POST.get('username')
            password = request.POST.get('password')
            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                logging.info(f'login user "{username}"')
                return HttpResponsePermanentRedirect('/')
            else:
                login_form = LoginForm()
                return render(request, 'login.html', {'form': login_form, 'errors': 'Неверный пароль'})
    else:
        login_form = LoginForm()
        return render(request, 'login.html', {'form': login_form, 'errors': ''})


def profile(request, user_id):
    user_from_request = request.user
    if not user_from_request.is_authenticated:
        return HttpResponsePermanentRedirect('/login')
    user = User.objects.get(id=user_id)
    if user_from_request == user:
        context = get_user_profile(user_from_request)
        return render(request, 'profile.html', context=context)
    return render(request, 'not_user_profile.html')


def change_password_view(request):
    user = request.user
    if user.is_authenticated():
        if request.method == 'POST':
            form = ChangePasswordForm(request.POST)
            if form.is_valid():
                pwd = form.cleaned_data['password']
                conf_pwd = form.cleaned_data['confirm_password']
                if pwd == conf_pwd:
                    user.set_password(pwd)
                    user.save()
                    return render('request', 'change_password_complete.html')
        return render('request', 'change_password.html')
    return HttpResponsePermanentRedirect('/')


def exception404(request):
    # ,exception
    return render(request, '404.html', status=404)


class StaffViews:
    """Views for staff"""

    @staticmethod
    def staff_only(func):
        '''Проверка доступа'''
        def wrapper(*args, **kwargs):
            if args[0].user.is_staff:
                return func(*args, **kwargs)
            return HttpResponsePermanentRedirect('/')
        return wrapper

    @staticmethod
    @staff_only
    def index(request):
        return render(request, 'staff/index.html')

    @staticmethod
    @staff_only
    def issue_books_view(request):
        page = request.GET.get('page', 1)
        genre = request.GET.get('genre', None)
        genres = Genre.objects.all()
        if genre:
            genre_label = genres.filter(id=genre)[0].title
            page = get_books_with_pagination(page=page, genres=[genre])
        else:
            genre_label = 'Все книги'
            page = get_books_with_pagination(page=page)
        context = {'page': page,
                   'genres': genres,
                   'genre_label': genre_label,
                   'media': settings.STATIC_URL,
                   'mode': 'staff'}
        return render(request, 'catalog.html', context=context)

    @staticmethod
    @staff_only
    def issue_a_book_view(request, book_id):
        book = Book.objects.get(id=book_id)
        count = book.bookinstance_set.all().filter(status=STATUS_FREE).count()
        if request.method == 'POST':
            issue_a_book_form = IssueABookForm(request.POST)
            if issue_a_book_form.is_valid():
                user_id = issue_a_book_form.cleaned_data.get('borrower')
                issue_a_book(book_id, user_id)
                return HttpResponsePermanentRedirect(reverse('borrow_page'))
        else:
            issue_a_book_form = IssueABookForm()
            context = {'book': book, 'count': count, 'form': issue_a_book_form}
            return render(request, 'staff/issue_a_book.html', context=context)

    @staticmethod
    @staff_only
    def borrow_view(request):
        book_instances = BookInstance.objects.all().filter(status=STATUS_BORROW)
        today = datetime.date.today()
        context = {'book_instances': book_instances, 'today': today}
        return render(request, 'staff/borrow.html', context=context)

    @staticmethod
    @staff_only
    def borrow_one_book_view(request, book_id):
        book_instance = BookInstance.objects.get(id=book_id)

        if request.method == 'POST':
            staff_borrow_book(book_instance, request.POST)
            return HttpResponsePermanentRedirect('/staff/borrow/')
        today = datetime.date.today()
        context = {'book_instance': book_instance, 'today': today}
        return render(request, 'staff/borrow_one_book.html', context=context)

    @staticmethod
    @staff_only
    def borrow_textbook_view(request):
        textbook_instances = TextbookInstance.objects.all().filter(status=STATUS_BORROW)
        context = {'textbook_instances': textbook_instances}
        return render(request, 'staff/borrow_textbook.html', context=context)

    @staticmethod
    @staff_only
    def add_book_view(request):
        if request.method == 'POST':
            add_book_form = AddNewBookForm(request.POST)
            if add_book_form.is_valid():
                title = add_book_form.cleaned_data.get('title')
                authors = add_book_form.cleaned_data.get('authors')
                genre = add_book_form.cleaned_data.get('genre')
                image = add_book_form.cleaned_data.get('image')
                publishing_house = add_book_form.cleaned_data.get('publishing_house')
                year_of_publication = add_book_form.cleaned_data.get('year_of_publication')
                count = add_book_form.cleaned_data.get('count')

                add_book_with_instances(title, authors, genre, image, publishing_house,
                                                 year_of_publication,
                                                 count)
                return render(request, 'staff/add_book_success.html')
        add_book_form = AddNewBookForm()
        return render(request, 'staff/add_book.html', {'form': add_book_form, 'errors': ''})

    @staticmethod
    @staff_only
    def add_book_instance_view(request):
        if request.method == 'POST':
            add_book_form = AddNewBookInstanceForm(request.POST)
            if add_book_form.is_valid():
                count = int(request.POST.get('count'))
                book_id = int(request.POST.get('book'))
                book = Book.objects.get(id=book_id)
                add_instances_to_book(book, count)
                return render(request, 'staff/add_book_success.html')
        else:
            add_book_form = AddNewBookInstanceForm()
            return render(request, 'staff/add_book_ins.html', {'form': add_book_form, 'errors': ''})

    @staticmethod
    @staff_only
    def add_textbooks_from_excel_view(request):  # отложено до лучших времен
        if request.method == 'POST':
            form = AddTextBookFromExcelForm(request.POST, request.FILES)
            file = request.FILES['file'].file
            wb = openpyxl.load_workbook(file)
            worksheet = wb.active
            for row in worksheet.iter_rows(0, worksheet.max_row):
                title = row[0].value
                authors = row[1].value
                genre = row[2].value
                pb_house = row[3].value
                year_of_publication = row[4].value
                # services.add_books_from_excel()
                print([title, authors, genre, pb_house, year_of_publication])
            return render(request, 'staff/add_book_success.html')
        form = AddTextBookFromExcelForm()
        return render(request, 'staff/add_textbooks_from_excel.html', context={'form': form})

    @staticmethod
    @staff_only
    def issue_textbooks_view(request):
        if request.method == 'POST':
            form = IssueTextbookForm(request.POST)
            if form.is_valid():
                textbook = form.cleaned_data['textbook']
                group = form.cleaned_data['group']
                borrower = form.cleaned_data['borrower']
                try:
                    issue_textbooks(textbook, group, borrower)
                except NotEnoughTextbooksError:
                    return
                return HttpResponsePermanentRedirect('/')
            return render(request, 'staff/issue_textbooks.html', context={'text': 'error', 'form': form})
        form = IssueTextbookForm()
        return render(request, 'staff/issue_textbooks.html', context={'text': 'text', 'form': form})

    @staticmethod
    @staff_only
    def issue_textbooks_list_view(request):
        textbooks = IssueTextbooks.objects.all()
        return render(request, 'staff/issued_textbook_list.html', context={'textbooks': textbooks})
